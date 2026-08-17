#!/usr/bin/env python3
"""將 Excel 規程清單與 PDF 純文字匯入 Policy Center PostgreSQL 草稿。

使用範例：
  python tools/import_policies.py --create-template policy_import.xlsx
  python tools/import_policies.py --excel policy_import.xlsx --pdf-dir ./pdfs
  python tools/import_policies.py --excel policy_import.xlsx --pdf-dir ./pdfs --apply

預設為預覽模式；只有加上 --apply 才會寫入資料庫。所有匯入項目都建立為
「草稿」，不會直接發布、不會取代現有規程，也不會繞過承認流程。
PDF 的圖片與表格結構均不匯入；僅保留一般段落／條文文字。
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import psycopg
from pypdf import PdfReader


REQUIRED_COLUMNS = ["policy_code", "category_code", "title_zh"]
TEMPLATE_COLUMNS = [
    "policy_code", "category_code", "title_zh", "content_zh", "pdf_zh",
    "title_ja", "content_ja", "pdf_ja", "effective_date", "revision_date",
    "revision_content", "revision_records", "revision_reason", "scheduled_publish_date", "created_by",
]
LANGUAGES = (("zh", "zh-TW"), ("ja", "ja-JP"))
CODE_PATTERN = re.compile(r"^DHT\d{1,2}-\d{4}$")
# PDF 文字擷取後，表格通常會保留 Tab、直線或連續欄位空白；此類列不匯入。
# 一般段落與條文會完整保留。無法從 PDF 文字層可靠辨識的表格列仍可能被當作文字，
# 這時請在匯入後的草稿內刪除該列。
TABLE_LIKE_PDF_LINE = re.compile(r"\t|\|| {3,}")


def clean(value: Any) -> str:
    """將 Excel 空白儲存格和非字串值統一成去除前後空白的文字。"""
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).strip()


def nullable_date(value: str) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError as error:
        raise ValueError(f"日期格式必須為 YYYY-MM-DD，目前是「{value}」。") from error


def revision_records_from_row(row: dict[str, str]) -> list[dict[str, str]]:
    """讀取多筆改訂紀錄；每行使用「YYYY-MM-DD｜改訂內容」格式。"""
    records: list[dict[str, str]] = []
    raw_records = row.get("revision_records", "")
    if raw_records:
        for line_number, line in enumerate(raw_records.splitlines(), start=1):
            line = line.strip()
            if not line:
                continue
            if "｜" not in line:
                raise ValueError(f"revision_records 第 {line_number} 行必須為「YYYY-MM-DD｜改訂內容」。")
            revision_date, content = (part.strip() for part in line.split("｜", 1))
            nullable_date(revision_date)
            if not content:
                raise ValueError(f"revision_records 第 {line_number} 行缺少改訂內容。")
            records.append({"date": revision_date, "content": content})
    elif row.get("revision_date") or row.get("revision_content"):
        revision_date = row.get("revision_date", "")
        nullable_date(revision_date)
        records.append({"date": revision_date, "content": row.get("revision_content", "")})
    return records


def extract_pdf_text(pdf_path: Path) -> str:
    """擷取文字型 PDF 的一般文字，不讀取圖片並略過可辨識的表格列。"""
    try:
        reader = PdfReader(str(pdf_path))
        pages = []
        for page in reader.pages:
            # pypdf 預設只取文字層；不讀取 page.images，因此圖片不會被帶入規程。
            lines = (page.extract_text() or "").splitlines()
            text_lines = [line.rstrip() for line in lines if not TABLE_LIKE_PDF_LINE.search(line)]
            pages.append("\n".join(text_lines).strip())
        text = "\n\n".join(page for page in pages if page).strip()
    except Exception as error:  # pypdf 的錯誤類型因 PDF 格式而異。
        raise ValueError(f"無法讀取 PDF：{pdf_path}（{error}）") from error
    if not text:
        raise ValueError(f"PDF 沒有可擷取的文字：{pdf_path}。若是掃描檔，請先進行 OCR。")
    return text


def resolve_pdf(value: str, pdf_dir: Path | None, excel_dir: Path) -> Path | None:
    """優先使用 Excel 所填路徑；相對路徑先從 --pdf-dir、再從 Excel 位置找。"""
    if not value:
        return None
    candidate = Path(value)
    options = [candidate] if candidate.is_absolute() else [
        pdf_dir / candidate if pdf_dir else candidate,
        excel_dir / candidate,
    ]
    for option in options:
        if option.exists() and option.is_file():
            return option
    raise ValueError(f"找不到 PDF：「{value}」。")


def read_rows(excel_path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = workbook.active
    headers = [clean(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"Excel 缺少必要欄位：{', '.join(missing)}")
    rows: list[dict[str, str]] = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {headers[index]: clean(value) for index, value in enumerate(cells) if index < len(headers) and headers[index]}
        if not any(row.values()):
            continue
        row["_row_number"] = str(row_number)
        rows.append(row)
    return rows


def make_translations(row: dict[str, str], pdf_dir: Path | None, excel_dir: Path) -> list[dict[str, Any]]:
    translations: list[dict[str, Any]] = []
    for short_name, language in LANGUAGES:
        title = row.get(f"title_{short_name}", "")
        content = row.get(f"content_{short_name}", "")
        pdf = resolve_pdf(row.get(f"pdf_{short_name}", ""), pdf_dir, excel_dir)
        if pdf:
            content = extract_pdf_text(pdf)
        if title:
            translations.append({
                "language": language,
                "title": title,
                "summary": "",
                "content": content,
                "chapters": [],
                # 匯入工具刻意不建立網頁表格，也不匯入 PDF／Excel 圖片。
                "tables": [],
                "images": [],
            })
        elif content or pdf:
            raise ValueError(f"{language} 有內文或 PDF，但未填 title_{short_name}。")
    if not translations:
        raise ValueError("至少需要一個語言的標題與內容。")
    return translations


def validate_row(row: dict[str, str], translations: list[dict[str, Any]]) -> None:
    code = row.get("policy_code", "")
    if not CODE_PATTERN.fullmatch(code):
        raise ValueError(f"policy_code 必須是 DHT2-0001 格式，目前是「{code}」。")
    if not row.get("category_code"):
        raise ValueError("category_code 不可空白。")
    if not any(item["language"] == "zh-TW" for item in translations):
        raise ValueError("至少需要中文標題 title_zh。")
    nullable_date(row.get("effective_date", ""))
    nullable_date(row.get("revision_date", ""))
    revision_records_from_row(row)
    nullable_date(row.get("scheduled_publish_date", ""))


def insert_policy(connection: psycopg.Connection, row: dict[str, str], translations: list[dict[str, Any]], default_user: str) -> None:
    """以單一交易建立 policy、草稿 request 與所有中日文內容。"""
    code = row["policy_code"]
    employee_no = row.get("created_by") or default_user
    revision_records = revision_records_from_row(row)
    with connection.transaction():
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1 FROM users WHERE employee_no = %s", (employee_no,))
            if cursor.fetchone() is None:
                raise ValueError(f"created_by「{employee_no}」不存在於 users 資料表。")
            cursor.execute("SELECT 1 FROM policy_categories WHERE category_code = %s", (row["category_code"],))
            if cursor.fetchone() is None:
                raise ValueError(f"category_code「{row['category_code']}」不存在。")
            cursor.execute("SELECT 1 FROM policies WHERE policy_code = %s", (code,))
            if cursor.fetchone() is not None:
                raise ValueError(f"規程編號「{code}」已存在；匯入工具不會覆蓋既有資料。")

            cursor.execute(
                """INSERT INTO policies (policy_code, category_code, status, effective_date, created_by)
                   VALUES (%s, %s, 'disabled', %s, %s)""",
                (code, row["category_code"], nullable_date(row.get("effective_date", "")), employee_no),
            )
            cursor.execute(
                """INSERT INTO policy_change_requests
                   (policy_code, change_kind, status, revision_reason, revision_date, revision_content, revision_records,
                    requested_effective_date, scheduled_publish_date, requires_approval, created_by)
                   VALUES (%s, 'new_policy', 'draft', %s, %s, %s, %s::jsonb, %s, %s, true, %s)
                   RETURNING change_request_id""",
                (
                    code,
                    row.get("revision_reason", ""),
                    nullable_date(revision_records[0]["date"]) if revision_records and revision_records[0]["date"] else None,
                    revision_records[0]["content"] if revision_records else "",
                    json.dumps(revision_records),
                    nullable_date(row.get("effective_date", "")),
                    nullable_date(row.get("scheduled_publish_date", "")),
                    employee_no,
                ),
            )
            change_request_id = cursor.fetchone()[0]
            for item in translations:
                cursor.execute(
                    """INSERT INTO policy_change_translations
                       (change_request_id, language, title, summary, content, chapters, tables, images)
                       VALUES (%s, %s, %s, %s, %s, %s::jsonb, %s::jsonb, %s::jsonb)""",
                    (change_request_id, item["language"], item["title"], item["summary"], item["content"], json.dumps(item["chapters"]), json.dumps(item["tables"]), json.dumps(item["images"])),
                )
            cursor.execute(
                """INSERT INTO policy_audit_logs
                   (actor_employee_no, policy_code, change_request_id, action, changed_fields, after_content, comment)
                   VALUES (%s, %s, %s, 'created', %s::jsonb, %s::jsonb, %s)""",
                (employee_no, code, change_request_id, json.dumps(["imported_excel", "translations", "revision_records"]), json.dumps({"translations": translations, "revisionRecords": revision_records}), row.get("revision_reason") or None),
            )


def create_template(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "規程匯入"
    sheet.append(TEMPLATE_COLUMNS)
    sheet.append([
        "DHT2-0001", "hr", "就業規程", "", "DHT2-0001_zh.pdf",
        "就業規則", "", "DHT2-0001_ja.pdf", "2026-09-01", "2026-09-01",
        "新制定", "2026-09-01｜新制定\n2026-10-01｜第 2 條文字修正", "新規程初版", "", "A0001",
    ])
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 24
    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="將 Excel + PDF 匯入 PostgreSQL 規程草稿")
    parser.add_argument("--excel", type=Path, help="規程 Excel 檔（.xlsx）")
    parser.add_argument("--pdf-dir", type=Path, help="Excel 中 PDF 相對路徑的基準資料夾")
    parser.add_argument("--database-url", default=os.environ.get("DATABASE_URL"), help="PostgreSQL DATABASE_URL；未填則讀取環境變數")
    parser.add_argument("--created-by", default="A0001", help="未填 created_by 時使用的員編，預設 A0001")
    parser.add_argument("--apply", action="store_true", help="實際寫入資料庫；未填時只檢查並預覽")
    parser.add_argument("--create-template", type=Path, help="建立 Excel 範本後結束")
    args = parser.parse_args()

    if args.create_template:
        create_template(args.create_template)
        print(f"已建立範本：{args.create_template}")
        return 0
    if not args.excel:
        parser.error("請提供 --excel，或使用 --create-template 建立範本。")
    if not args.excel.exists():
        parser.error(f"找不到 Excel：{args.excel}")

    try:
        rows = read_rows(args.excel)
        prepared: list[tuple[dict[str, str], list[dict[str, Any]]]] = []
        for row in rows:
            translations = make_translations(row, args.pdf_dir, args.excel.parent)
            validate_row(row, translations)
            prepared.append((row, translations))
        if not prepared:
            raise ValueError("Excel 沒有可匯入的資料列。")
    except ValueError as error:
        print(f"匯入檢查失敗：{error}", file=sys.stderr)
        return 1

    for row, translations in prepared:
        languages = ", ".join(item["language"] for item in translations)
        print(f"第 {row['_row_number']} 列：{row['policy_code']} / {row['category_code']} / {languages}")
    if not args.apply:
        print(f"\n以上 {len(prepared)} 筆檢查完成。這是預覽，尚未寫入資料庫；確認後加上 --apply。")
        return 0
    if not args.database_url:
        print("缺少 DATABASE_URL。請先設定環境變數或傳入 --database-url。", file=sys.stderr)
        return 1

    try:
        with psycopg.connect(args.database_url, options="-c search_path=role_web,public") as connection:
            for row, translations in prepared:
                insert_policy(connection, row, translations, args.created_by)
                print(f"已建立草稿：{row['policy_code']}")
    except (psycopg.Error, ValueError) as error:
        print(f"匯入失敗：{error}", file=sys.stderr)
        return 1
    print(f"\n完成：共建立 {len(prepared)} 筆草稿。請以 Admin 登入後送交承認。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
